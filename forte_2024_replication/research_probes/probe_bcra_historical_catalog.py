"""Inspect the small BCRA historical-series catalogue without downloading data.

The probe verifies catalogue mnemonics and prints only rows matching concepts
needed by the Forte (2024) audit. It does not download the complete histories.
"""

from __future__ import annotations

import argparse
import io
import json

import requests
from openpyxl import load_workbook


CATALOG_URL = (
    "https://www.bcra.gob.ar/archivos/Pdfs/PublicacionesEstadisticas/"
    "Series_estadisticas.xlsx"
)
TERMS = (
    "m2",
    "base monetaria",
    "crédito total",
    "credito total",
    "crédito en pesos",
    "credito en pesos",
    "préstamos totales",
    "prestamos totales",
    "total de préstamos",
    "total de prestamos",
    "hasta 60 días",
    "hasta 60 dias",
)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--term",
        action="append",
        help="Case-insensitive catalogue term; may be repeated.",
    )
    args = parser.parse_args()
    terms = tuple(term.lower() for term in args.term) if args.term else TERMS
    response = requests.get(CATALOG_URL, timeout=60)
    response.raise_for_status()
    workbook = load_workbook(
        io.BytesIO(response.content), read_only=True, data_only=True
    )
    for sheet in workbook.worksheets:
        rows = sheet.iter_rows(values_only=True)
        header = tuple(next(rows))
        for row in rows:
            record = dict(zip(header, row))
            description_keys = [
                key for key in record if str(key).startswith("tx_serie")
            ]
            if not description_keys:
                continue
            description_key = description_keys[0]
            searchable = " | ".join(
                str(record.get(key) or "")
                for key in (description_key, "nombre")
            ).lower()
            if any(term in searchable for term in terms):
                compact = {
                    key: record.get(key)
                    for key in (
                        "cd_serie",
                        description_key,
                        "periodicidad",
                        "nombre",
                        "tipo",
                        "unidad de expresión",
                        "moneda de denominación",
                        "sector",
                        "tipo de titular",
                        "Ubicación TXT",
                    )
                }
                compact["tx_nota_metodologica"] = str(
                    record.get("tx_nota_metodologica") or ""
                )[:500]
                print(
                    json.dumps(
                        {"sheet": sheet.title, "record": compact},
                        ensure_ascii=False,
                        default=str,
                    )
                )


if __name__ == "__main__":
    main()
