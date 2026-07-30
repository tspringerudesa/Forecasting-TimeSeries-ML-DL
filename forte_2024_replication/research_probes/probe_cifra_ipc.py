"""Verify the tiny public CIFRA IPC-Provincias workbook.

The workbook is about 15 KB. It is read in memory and discarded; only sheet
names, dimensions, and boundary rows are printed.
"""

from __future__ import annotations

import io
import json
from datetime import date, datetime

import requests
from openpyxl import load_workbook


URL = (
    "https://centrocifra.org.ar/wp-content/uploads/2023/08/"
    "IPC-Provincias-2007-2018.xlsx"
)


def main() -> None:
    response = requests.get(
        URL,
        headers={
            "User-Agent": "Mozilla/5.0 source-audit/0.1",
            "Referer": "https://centrocifra.org.ar/estadisticas/ipc-provincias/",
        },
        timeout=30,
    )
    response.raise_for_status()
    workbook = load_workbook(io.BytesIO(response.content), read_only=True, data_only=True)
    for sheet in workbook.worksheets:
        rows = list(sheet.iter_rows(values_only=True))
        nonempty = [row for row in rows if any(value is not None for value in row)]
        dated = [
            row
            for row in nonempty
            if row
            and (
                isinstance(row[0], (date, datetime))
                or (
                    isinstance(row[0], str)
                    and any(char.isdigit() for char in row[0])
                    and ("/" in row[0] or "-" in row[0])
                )
            )
        ]
        print(
            json.dumps(
                {
                    "response_bytes": len(response.content),
                    "sheet": sheet.title,
                    "rows": sheet.max_row,
                    "columns": sheet.max_column,
                    "first_nonempty_row": nonempty[0] if nonempty else None,
                    "last_nonempty_row": nonempty[-1] if nonempty else None,
                    "first_dated_row": dated[0] if dated else None,
                    "last_dated_row": dated[-1] if dated else None,
                },
                ensure_ascii=False,
                default=str,
            )
        )


if __name__ == "__main__":
    main()
